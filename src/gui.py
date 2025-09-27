import json
import os
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Set

import tkinter as tk
from tkinter import ttk, messagebox

from openpyxl import Workbook

from src import ebay_api


SEEN_IDS_FILE = Path(__file__).resolve().parent / "seen_ids.json"


def load_seen_ids() -> Set[str]:
    """Load the set of seen eBay item IDs from disk."""

    if not SEEN_IDS_FILE.exists():
        return set()

    try:
        data = json.loads(SEEN_IDS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return set()

    if isinstance(data, list):
        return {str(item_id) for item_id in data if str(item_id).strip()}

    return set()


def save_seen_ids(seen_ids: Set[str]) -> None:
    """Persist the seen item IDs to disk."""

    try:
        SEEN_IDS_FILE.write_text(
            json.dumps(sorted(seen_ids)), encoding="utf-8"
        )
    except OSError:
        # Persisting seen IDs should not break the user experience if it fails.
        pass


def deduplicate_results(
    results: List[dict], allow_seen: bool, seen_ids: Set[str]
) -> List[dict]:
    """Filter out results that have already been seen unless allowed."""

    filtered_results: List[dict] = []

    for listing in results:
        item_id = str(listing.get("item_id") or "").strip()
        already_seen = item_id in seen_ids if item_id else False

        if allow_seen or not already_seen or not item_id:
            filtered_results.append(listing)

        if item_id:
            seen_ids.add(item_id)

    return filtered_results


session_token: Optional[str] = None


def start_token_window() -> None:
    """Launch the token entry window and transition to the main GUI."""

    token_window = tk.Tk()
    token_window.title("Enter eBay OAuth Token")

    tk.Label(token_window, text="Enter eBay OAuth Token").pack(pady=(15, 5), padx=15)

    entry = tk.Entry(token_window, width=60)
    entry.pack(padx=15)
    entry.focus_set()

    def submit_token() -> None:
        global session_token
        token = entry.get().strip()
        if not token:
            messagebox.showerror("Error", "Token cannot be empty")
            return
        session_token = token
        print("Token entered and saved")
        token_window.destroy()
        launch_main_window()

    ttk.Button(token_window, text="Submit", command=submit_token).pack(pady=15)

    token_window.lift()
    token_window.attributes("-topmost", True)
    token_window.mainloop()


def _build_main_window(root: tk.Tk) -> None:
    """Main Watch Listings Fetcher window."""

    root.title("Watch Listings Fetcher")

    # Search fields
    tk.Label(root, text="Search Query").grid(row=0, column=0, sticky="w")
    search_query_entry = tk.Entry(root)
    search_query_entry.grid(row=0, column=1)

    tk.Label(root, text="Brand").grid(row=1, column=0, sticky="w")
    brand_entry = tk.Entry(root)
    brand_entry.grid(row=1, column=1)

    tk.Label(root, text="Model").grid(row=2, column=0, sticky="w")
    model_entry = tk.Entry(root)
    model_entry.grid(row=2, column=1)

    tk.Label(root, text="Min Price").grid(row=3, column=0, sticky="w")
    min_price_entry = tk.Entry(root)
    min_price_entry.grid(row=3, column=1)

    tk.Label(root, text="Max Price").grid(row=4, column=0, sticky="w")
    max_price_entry = tk.Entry(root)
    max_price_entry.grid(row=4, column=1)

    tk.Label(root, text="Condition").grid(row=5, column=0, sticky="w")
    condition_entry = tk.Entry(root)
    condition_entry.grid(row=5, column=1)

    tk.Label(root, text="Listing Type").grid(row=6, column=0, sticky="w")
    listing_type_entry = tk.Entry(root)
    listing_type_entry.grid(row=6, column=1)

    tk.Label(root, text="Auction Only").grid(row=7, column=0, sticky="w")
    auction_only_var = tk.BooleanVar(value=False)
    tk.Checkbutton(root, text="Auction Only", variable=auction_only_var).grid(
        row=7, column=1, sticky="w"
    )

    tk.Label(root, text="Max Time Left").grid(row=8, column=0, sticky="w")
    max_time_entry = tk.Entry(root)
    max_time_entry.grid(row=8, column=1)

    tk.Label(root, text="Exclude Keywords").grid(row=9, column=0, sticky="w")
    exclude_entry = tk.Entry(root)
    exclude_entry.grid(row=9, column=1)

    tk.Label(root, text="Results").grid(row=10, column=0, sticky="w")
    results_entry = tk.Entry(root)
    results_entry.insert(0, "20")
    results_entry.grid(row=10, column=1)

    show_seen_var = tk.BooleanVar(value=False)
    tk.Checkbutton(
        root,
        text="Show results you've seen before",
        variable=show_seen_var,
    ).grid(row=11, column=0, columnspan=2, sticky="w", pady=(5, 0))

    # Output box
    output = tk.Text(root, height=15, width=80)

    results_data: List[dict] = []

    def fetch_listings():
        if not session_token:
            messagebox.showerror("Error", "No token available. Please restart the application.")
            return
        query = {
            "search_query": search_query_entry.get().strip(),
            "brand": brand_entry.get().strip(),
            "model": model_entry.get().strip(),
            "min_price": min_price_entry.get().strip(),
            "max_price": max_price_entry.get().strip(),
            "condition": condition_entry.get().strip(),
            "listing_type": listing_type_entry.get().strip(),
            "auction_only": auction_only_var.get(),
            "max_time_left": max_time_entry.get().strip(),
            "exclude_keywords": exclude_entry.get().strip(),
            "limit": results_entry.get().strip(),
        }
        try:
            listings = ebay_api.fetch_listings(query, session_token)

            seen_ids = load_seen_ids()
            allow_seen = show_seen_var.get()
            filtered_listings = deduplicate_results(listings, allow_seen, seen_ids)
            save_seen_ids(seen_ids)

            results_data.clear()
            results_data.extend(filtered_listings)

            output.delete(1.0, tk.END)
            computed_query = query.get("computed_query")
            if computed_query:
                output.insert(tk.END, f"Search query: {computed_query}\n\n")
            else:
                output.insert(tk.END, "Search query: (none)\n\n")
            if not listings:
                output.insert(tk.END, "No listings found.\n")
            else:
                hidden_count = len(listings) - len(filtered_listings)
                if hidden_count and not allow_seen:
                    output.insert(
                        tk.END,
                        f"{hidden_count} seen listing(s) hidden. Check the option above to view them.\n\n",
                    )

                preview = filtered_listings[:5]
                if len(filtered_listings) > 5:
                    output.insert(
                        tk.END,
                        f"Showing top 5 of {len(filtered_listings)} results. Use Export Results for the full list.\n\n",
                    )
                elif filtered_listings:
                    output.insert(tk.END, f"Showing {len(preview)} result(s).\n\n")
                else:
                    output.insert(tk.END, "No unseen listings found.\n")

                for idx, item in enumerate(preview, start=1):
                    title = item.get("title", "No title")
                    price = item.get("price", "N/A")
                    output.insert(tk.END, f"{idx}. {title} - {price}\n")
        except Exception as e:
            output.delete(1.0, tk.END)
            output.insert(tk.END, f"Error: {e}\n")

    def export_results():
        if not results_data:
            messagebox.showinfo("No Results", "No listings to export. Please fetch listings first.")
            return

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Results"

            headers = ["Title", "Price", "Condition", "Time Left", "Seller Rating", "URL"]
            sheet.append(headers)

            for listing in results_data:
                row = [
                    listing.get("title", ""),
                    listing.get("price", ""),
                    listing.get("condition", ""),
                    listing.get("time_left", ""),
                    listing.get("seller_rating", ""),
                    listing.get("url", ""),
                ]
                sheet.append(row)
                url_cell = sheet.cell(row=sheet.max_row, column=6)
                url_value = listing.get("url")
                if url_value:
                    url_cell.hyperlink = url_value
                    url_cell.style = "Hyperlink"

            desktop_path = Path(os.path.expanduser("~")) / "Desktop"
            desktop_path.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            file_path = desktop_path / f"results_{timestamp}.xlsx"

            workbook.save(file_path)
            messagebox.showinfo(
                "Export Complete",
                f"Exported {len(results_data)} listings to {file_path}",
            )
        except Exception as exc:
            messagebox.showerror(
                "Export Failed", f"An error occurred while exporting results: {exc}"
            )

    ttk.Button(root, text="Fetch Listings", command=fetch_listings).grid(
        row=12, column=0, columnspan=2, pady=(10, 0)
    )
    ttk.Button(root, text="Export Results", command=export_results).grid(
        row=13, column=0, columnspan=2, pady=(10, 0)
    )
    output.grid(row=14, column=0, columnspan=2, pady=10)


def launch_main_window() -> None:
    root = tk.Tk()
    _build_main_window(root)
    root.mainloop()


def run() -> None:
    start_token_window()


if __name__ == "__main__":
    run()
